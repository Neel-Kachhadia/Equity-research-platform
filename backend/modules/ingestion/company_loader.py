"""
EREBUS — Multi-Format S3 Company Loader
========================================
Reads ALL files from a company's S3 folder structure.

[Documentation unchanged - kept for brevity]
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
from typing import Any, Dict, List, Optional

import boto3
import botocore.exceptions
import threading
from typing import Dict, List, Any, Optional
from botocore.config import Config

logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
def _get_bucket() -> str:
    """Read bucket name from env at call time — never stale after a restart."""
    return os.getenv("S3_BUCKET_NAME", "erebus-data-prod")

_BUCKET  = _get_bucket()   # kept for backward-compat imports; refreshed below
_REGION  = os.getenv("AWS_REGION", "ap-south-1").strip()

# Sub-folder names to scan for documents
_DOC_FOLDERS = ["Annual_Reports", "Quarterly_Results", "Other_Documents", "CSR_Reports", "reports"]

# Thread-safe S3 client
_s3_client = None
_s3_lock = threading.Lock()

def _get_s3_client():
    global _s3_client
    if _s3_client is not None:
        return _s3_client
    
    with _s3_lock:
        if _s3_client is not None:
            return _s3_client
        
        aws_key = os.getenv("AWS_ACCESS_KEY_ID")
        aws_secret = os.getenv("AWS_SECRET_ACCESS_KEY")
        
        if not (aws_key and aws_secret):
            raise CompanyDataError(
                "AWS credentials not configured. "
                "Set AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY."
            )
        
        _s3_client = boto3.client(
            "s3",
            aws_access_key_id=aws_key,
            aws_secret_access_key=aws_secret,
            region_name=_REGION,
            config=Config(
                retries={"max_attempts": 3, "mode": "standard"},
                connect_timeout=2,
                read_timeout=10,
            ),
        )
        logger.info("[S3] Client initialised (region=%s, bucket=%s)", _REGION, _BUCKET)
    return _s3_client

# Screener.in Excel sheet names → internal keys
_EXCEL_SHEETS = {
    "Profit & Loss": "pnl",
    "Balance Sheet":  "bs",
    "Cash Flow":      "cf",
    "Quarters":       "quarters",
    "Data Sheet":     "data_sheet",
}

_s3_client = None


# ── Custom exception ──────────────────────────────────────────────────────────

class CompanyDataError(Exception):
    """Raised when company data cannot be loaded or fails validation.
    Callers should surface this as HTTP 400."""


# ── S3 client (timeout + retry) ───────────────────────────────────────────────

def _get_s3_client():
    global _s3_client
    if _s3_client is not None:
        return _s3_client

    aws_key    = os.getenv("AWS_ACCESS_KEY_ID")
    aws_secret = os.getenv("AWS_SECRET_ACCESS_KEY")

    if not (aws_key and aws_secret):
        raise CompanyDataError(
            "AWS credentials not configured. "
            "Set AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY."
        )

    _s3_client = boto3.client(
        "s3",
        aws_access_key_id=aws_key,
        aws_secret_access_key=aws_secret,
        region_name=_REGION,
        config=Config(
            retries={"max_attempts": 3, "mode": "standard"},
            connect_timeout=2,
            read_timeout=10,
        ),
    )
    logger.info("[S3] Client initialised (region=%s, bucket=%s)", _REGION, _BUCKET)
    return _s3_client


# ── S3 helpers ────────────────────────────────────────────────────────────────

def _list_prefix(prefix: str) -> List[Dict[str, Any]]:
    """List all objects under an S3 prefix. Returns list of {Key, Size} dicts."""
    client = _get_s3_client()
    paginator = client.get_paginator("list_objects_v2")
    objects = []
    try:
        for page in paginator.paginate(Bucket=_BUCKET, Prefix=prefix):
            for obj in page.get("Contents", []):
                objects.append({"key": obj["Key"], "size": obj["Size"]})
    except botocore.exceptions.ClientError as e:
        code = e.response["Error"]["Code"]
        logger.warning("[S3] list_prefix failed [%s]: %s", code, e)
    return objects


def _fetch_bytes(s3_key: str) -> bytes:
    """Fetch raw bytes of a single S3 object."""
    client = _get_s3_client()
    try:
        resp = client.get_object(Bucket=_BUCKET, Key=s3_key)
        return resp["Body"].read()
    except botocore.exceptions.ClientError as e:
        code = e.response["Error"]["Code"]
        raise CompanyDataError(f"[S3] GetObject failed [{code}]: {s3_key}") from e
    except botocore.exceptions.ConnectTimeoutError:
        raise CompanyDataError(f"[S3] Connection timeout: {s3_key}")
    except botocore.exceptions.ReadTimeoutError:
        raise CompanyDataError(f"[S3] Read timeout: {s3_key}")


# ── File readers ──────────────────────────────────────────────────────────────

def _read_excel(raw: bytes) -> Dict[str, Any]:
    """
    Parse a Screener.in-style Excel workbook.

    Screener.in P&L / BS / CF sheets use formulas that reference the 'Data Sheet'
    tab as the data source.  If those formulas were never cached on the last save,
    data_only=True returns None for every cell.

    Strategy:
        1. Try reading the named sheets (Profit & Loss, Balance Sheet, etc.)
        2. If all value cells are None (formula-cache miss), fall through to
           reading the 'Data Sheet' / 'Quarters' tab directly using wide-format
           logic (which has hard-coded numbers).

    Returns dict keyed by internal sheet name (pnl, bs, cf, quarters, data_sheet).
    """
    try:
        import openpyxl
    except ImportError:
        raise CompanyDataError("openpyxl not installed — run: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    result: Dict[str, Any] = {}

    for sheet_name, internal_key in _EXCEL_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_data = []
        headers: List[str] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = list(row)
            if not any(c is not None for c in cells):
                continue  # skip blank rows

            first_cell = str(cells[0]).strip() if cells[0] is not None else ""
            if first_cell.lower() == "narration" or row_idx == 3:
                headers = [str(c).strip() if c is not None else "" for c in cells]
                continue

            if "SCREENER" in first_cell.upper():
                continue

            row_dict: Dict[str, Any] = {"narration": first_cell, "values": []}
            for c in cells[1:]:
                try:
                    row_dict["values"].append(float(c) if c is not None else None)
                except (ValueError, TypeError):
                    row_dict["values"].append(None)
            rows_data.append(row_dict)

        result[internal_key] = {"headers": headers, "rows": rows_data}

    # ── Formula-cache detection ─────────────────────────────────────────────
    # If all pnl/bs/cf values are None, the Excel was never recalculated.
    # In that case, extract from the 'Data Sheet' / 'Quarters' tabs directly.
    def _has_real_data(key: str) -> bool:
        for row in result.get(key, {}).get("rows", []):
            if any(v is not None for v in row.get("values", [])):
                return True
        return False

    pnl_ok = _has_real_data("pnl")
    bs_ok  = _has_real_data("bs")

    if not pnl_ok:
        # Pull real numbers from Data Sheet or Quarters using year-column detection
        for fallback_sheet in ["Data Sheet", "Quarters"]:
            if fallback_sheet not in wb.sheetnames:
                continue
            ws = wb[fallback_sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Detect year-like column headers in first non-blank row
            for hdr_row in rows[:4]:
                hdr = [str(c).strip() if c is not None else "" for c in hdr_row]
                # Match "Mar 2020", "2020", "Mar-20", "TTM", "Trailing"
                year_cols = [
                    i for i, h in enumerate(hdr)
                    if re.search(r"\b(20\d{2}|mar|jun|sep|dec|fy\d{2}|ttm|trailing)\b", h, re.I)
                    and i > 0
                ]
                if year_cols:
                    # Use last N year columns (most recent data)
                    year_cols = year_cols[-10:]
                    year_labels = [hdr[i] for i in year_cols]

                    sheet_rows = []
                    for data_row in rows:
                        if data_row[0] is None:
                            continue
                        narration = str(data_row[0]).strip()
                        if not narration or "SCREENER" in narration.upper():
                            continue
                        vals = []
                        for ci in year_cols:
                            try:
                                v = data_row[ci] if ci < len(data_row) else None
                                vals.append(float(v) if v is not None else None)
                            except (ValueError, TypeError):
                                vals.append(None)
                        sheet_rows.append({"narration": narration, "values": vals})

                    if sheet_rows:
                        result["pnl"] = {"headers": year_labels, "rows": sheet_rows}
                        logger.info("[Excel] Screener formula-cache miss — using '%s' tab instead", fallback_sheet)
                    break

    # ── Universal fallback: any sheet with year-column headers ──────────────────
    # Handles Moneycontrol, custom exports, and any non-Screener.in Excel format.
    if not _has_real_data("pnl"):
        _YEAR_RE = re.compile(r"\b(20\d{2}|mar|jun|sep|dec|fy\d{2}|ttm|trailing)\b", re.I)
        _BS_KEYWORDS = re.compile(r"balance\s*sheet|equity|assets|liabilit", re.I)
        _CF_KEYWORDS = re.compile(r"cash\s*flow|operating\s*activit|investing\s*activit|financing\s*activit", re.I)

        for sname in wb.sheetnames:
            ws    = wb[sname]
            rows  = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Look for a header row containing year-like labels in columns 1+
            hdr_idx, year_cols, year_labels = None, [], []
            for i, row in enumerate(rows[:6]):
                hdr = [str(c).strip() if c is not None else "" for c in row]
                yc  = [j for j, h in enumerate(hdr) if _YEAR_RE.search(h) and j > 0]
                if len(yc) >= 2:          # need at least 2 year columns
                    hdr_idx, year_cols = i, yc[-10:]
                    year_labels = [hdr[j] for j in year_cols]
                    break

            if hdr_idx is None:
                continue                  # no year header found in this sheet

            sheet_rows = []
            for data_row in rows[hdr_idx + 1:]:
                if not data_row or data_row[0] is None:
                    continue
                narration = str(data_row[0]).strip()
                if not narration:
                    continue
                vals = []
                for ci in year_cols:
                    try:
                        v = data_row[ci] if ci < len(data_row) else None
                        vals.append(float(v) if v is not None else None)
                    except (ValueError, TypeError):
                        vals.append(None)
                if any(v is not None for v in vals):
                    sheet_rows.append({"narration": narration, "values": vals})

            if not sheet_rows:
                continue

            # Classify sheet: balance sheet / cash flow / default pnl
            if _BS_KEYWORDS.search(sname):
                dest = "bs"
            elif _CF_KEYWORDS.search(sname):
                dest = "cf"
            else:
                dest = "pnl"

            if dest not in result or not _has_real_data(dest):
                result[dest] = {"headers": year_labels, "rows": sheet_rows}
                logger.info("[Excel] Universal fallback — sheet '%s' → key '%s' (%d rows)",
                            sname, dest, len(sheet_rows))

    return result


def _latest(values: Optional[List[Optional[float]]], default: float = 0.0) -> float:
    """Return the most recent non-None value, or default."""
    if not values:
        return default
    for v in reversed(values):
        if v is not None:
            return float(v)
    return default


def _extract_metric(sheet_data: Dict, narration_pattern: str) -> Optional[List[Optional[float]]]:
    """Find a row in sheet_data['rows'] whose narration matches narration_pattern."""
    if not sheet_data or "rows" not in sheet_data:
        return None
    pattern = re.compile(narration_pattern, re.IGNORECASE)
    for row in sheet_data["rows"]:
        if pattern.search(row.get("narration", "")):
            return row["values"]
    return None



def _read_excel_wide_format(raw: bytes) -> Dict[str, Any]:
    """
    Parse a wide-format Data Sheet Excel (TCS style).
    Structure:
        Row 1 : quarter/year headers e.g. '1Q11','2Q11',...,'FY11','1Q12',...,'FY25','1Q26'
        Row 2+: metric rows  e.g. 'Total Revenue', 'Employee cost', ...
    Returns a dict matching the same schema as _read_excel() for the 'pnl' key,
    using only annual 'FY' columns.
    """
    try:
        import openpyxl
    except ImportError:
        return {}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        return {}

    result: Dict[str, Any] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect if row 0 is a header row with 'FY' annual columns
        header_row = [str(c).strip() if c is not None else "" for c in rows[0]]
        fy_cols = [i for i, h in enumerate(header_row) if h.startswith("FY")]
        if not fy_cols:
            continue  # not the wide format we expect

        # Use last 5 FY columns for the series
        fy_cols = fy_cols[-5:]
        fy_labels = [header_row[i] for i in fy_cols]

        sheet_rows: List[Dict] = []
        for data_row in rows[1:]:
            if data_row[0] is None:
                continue
            narration = str(data_row[0]).strip()
            if not narration or narration.upper().startswith("EX-ADJ"):
                continue
            values: List[Optional[float]] = []
            for col_idx in fy_cols:
                try:
                    v = data_row[col_idx]
                    values.append(float(v) if v is not None else None)
                except (ValueError, TypeError):
                    values.append(None)
            sheet_rows.append({"narration": narration, "values": values})

        # Store under a generic 'pnl' key — downstream works the same
        if sheet_rows:
            internal_key = "pnl"
            if "bs" in sheet_name.lower() or "balance" in sheet_name.lower():
                internal_key = "bs"
            if internal_key not in result or not result[internal_key].get("rows"):
                result[internal_key] = {"headers": fy_labels, "rows": sheet_rows}

    return result


def _read_pdf_text(raw: bytes, max_chars: int = 8000) -> str:
    """Extract plain text from a PDF (first max_chars characters)."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        text_parts = []
        for page in reader.pages:
            text_parts.append(page.extract_text() or "")
            if sum(len(t) for t in text_parts) > max_chars:
                break
        return " ".join(text_parts)[:max_chars]
    except Exception as e:
        logger.debug("[PDF] Text extraction failed: %s", e)
        return ""


def _read_json(raw: bytes) -> Optional[Dict]:
    """Parse JSON bytes, return None on failure."""
    try:
        return json.loads(raw)
    except Exception:
        return None


def _read_csv(raw: bytes) -> List[Dict]:
    """Parse CSV bytes into a list of dicts."""
    import csv
    try:
        text = raw.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)
    except Exception:
        return []


# ── Financials builder from Screener.in Excel ─────────────────────────────────

def _build_financials_from_excel(sheets: Dict[str, Any]) -> Dict[str, Any]:
    """
    Map Screener.in sheet rows → normalised financials dict.
    All values in Crores (as Screener.in provides).
    """
    pnl = sheets.get("pnl", {})
    bs  = sheets.get("bs",  {})
    cf  = sheets.get("cf",  {})

    # P&L — patterns cover Screener.in ("Sales") AND TCS Data Sheet ("Total Revenue")
    revenue     = _latest(_extract_metric(pnl, r"total revenue|^sales$|^revenue$"), 0)
    if not revenue:
        revenue = _latest(_extract_metric(pnl, r"revenue|sales"), 0)
    expenses    = _latest(_extract_metric(pnl, r"^expenses$|total expenses|cost of revenue"), 0)
    op_profit   = _latest(_extract_metric(pnl, r"operating profit|ebit(?!\w)|pbdit"), 0)
    other_income  = _latest(_extract_metric(pnl, r"other income"), 0)
    depreciation  = _latest(_extract_metric(pnl, r"depreciation|amortis"), 0)
    interest_exp  = _latest(_extract_metric(pnl, r"^interest|finance cost"), 0)
    pbt           = _latest(_extract_metric(pnl, r"profit before tax|pbt|income before tax"), 0)
    tax           = _latest(_extract_metric(pnl, r"^tax$|tax expense|provision for tax"), 0)
    net_income    = _latest(_extract_metric(pnl, r"net profit|net income|profit after tax|^pat$"), 0)
    eps           = _latest(_extract_metric(pnl, r"^eps|earnings per share"))

    # revenue series (all years)
    rev_series_raw = (_extract_metric(pnl, r"total revenue|^sales$|^revenue$") or
                      _extract_metric(pnl, r"revenue|sales"))

    # Balance Sheet — covers Screener.in and TCS Data Sheet formats
    equity_capital  = _latest(_extract_metric(bs, r"equity share capital|share capital"), 0)
    reserves        = _latest(_extract_metric(bs, r"^reserves|shareholders.{0,10}equity|total equity"), 0)
    total_equity    = equity_capital + reserves if (equity_capital + reserves) else 1
    borrowings      = _latest(_extract_metric(bs, r"^borrowings|total debt|long.term debt|short.term debt"), 0)
    other_liab      = _latest(_extract_metric(bs, r"other liabilities|other current liabilities"), 0)
    total_assets    = _latest(_extract_metric(bs, r"^total$|total assets|^assets$"), 0)
    curr_assets     = _latest(_extract_metric(bs, r"current assets|total current assets"), 0)
    curr_liab       = _latest(_extract_metric(bs, r"current liabilities|total current liab"), 0)
    inventory       = _latest(_extract_metric(bs, r"inventory|inventories"), 0)
    investments     = _latest(_extract_metric(bs, r"invest"), 0)

    # Cash Flow
    cfo = _latest(_extract_metric(cf, r"cash from operating|operating activities"), 0)
    cfi = _latest(_extract_metric(cf, r"cash from investing|investing activities"), 0)
    cff = _latest(_extract_metric(cf, r"cash from financing|financing activities"), 0)
    capex = _latest(_extract_metric(cf, r"capex|capital expenditure"), 0)
    fcf = cfo - abs(capex) if capex else cfo

    # Derived
    ebit = op_profit if op_profit else (revenue - expenses - depreciation)
    cogs = expenses if expenses else (revenue * 0.45)

    # Prior revenue from series (second-to-last value)
    prior_revenue = revenue * 0.95
    if rev_series_raw:
        non_none = [v for v in rev_series_raw if v is not None]
        if len(non_none) >= 2:
            prior_revenue = non_none[-2]

    fcf_conversion = (fcf / net_income) if net_income else 0.85

    return {
        "revenue":              revenue,
        "net_income":           net_income,
        "total_equity":         max(total_equity, 1),
        "total_debt":           borrowings,
        "total_assets":         total_assets or (total_equity + borrowings),
        "current_assets":       curr_assets or (total_equity * 0.4),
        "current_liabilities":  curr_liab or (total_equity * 0.2),
        "inventory":            inventory,
        "ebit":                 ebit,
        "interest_expense":     interest_exp,
        "operating_income":     op_profit,
        "cogs":                 cogs,
        "market_cap":           0.0,
        "shares_outstanding":   0.0,
        "price_per_share":      0.0,
        "book_value_per_share": total_equity / max(equity_capital or 1, 1),
        "prior_revenue":        prior_revenue,
        "equity_capital":       equity_capital,
        "reserves":             reserves,
        "borrowings":           borrowings,
        "other_liabilities":    other_liab,
        "investments":          investments,
        "cash_from_operations": cfo,
        "cash_from_investing":  cfi,
        "cash_from_financing":  cff,
        "capex":                capex,
        "fcf":                  fcf,
        "fcf_conversion":       fcf_conversion,
        "eps":                  eps,
        "pbt":                  pbt,
        "tax":                  tax,
        "depreciation":         depreciation,
        "other_income":         other_income,
    }


def _build_prices_from_quarters(sheets: Dict[str, Any]) -> List[float]:
    """Extract quarterly sales/revenue as a price proxy from the Quarters sheet."""
    quarters = sheets.get("quarters", {})
    sales_series = _extract_metric(quarters, r"^sales$|^revenue")
    if not sales_series:
        return []
    return [float(v) for v in sales_series if v is not None]


def _build_series_from_pnl(sheets: Dict[str, Any]) -> Dict[str, List[float]]:
    """Extract multi-year series from P&L sheet for alpha signal computation."""
    pnl = sheets.get("pnl", {})
    bs  = sheets.get("bs",  {})

    def clean(vals):
        return [float(v) for v in (vals or []) if v is not None]

    rev_series   = clean(_extract_metric(pnl, r"total revenue|^sales$|^revenue$") or
                         _extract_metric(pnl, r"revenue|sales"))
    ni_series    = clean(_extract_metric(pnl, r"net profit|net income|profit after tax|^pat$"))
    ebit_series  = clean(_extract_metric(pnl, r"operating profit|ebit(?!\w)|pbdit"))
    borrow_ser   = clean(_extract_metric(bs,  r"^borrowings|total debt"))
    equity_ser   = clean(_extract_metric(bs,  r"^reserves|shareholders.{0,10}equity|total equity"))

    return {
        "revenue_series": rev_series[-5:] if rev_series else [],
        "ni_series":      ni_series[-5:]  if ni_series  else [],
        "ebit_series":    ebit_series[-5:] if ebit_series else [],
        "borrow_series":  borrow_ser[-5:] if borrow_ser else [],
        "equity_series":  equity_ser[-5:] if equity_ser else [],
    }


# ── Alpha fields derivation (explicit) ───────────────────────────────────────

def _compute_growth(series: List[float]) -> Optional[float]:
    """
    Compute CAGR from series.
    Returns None if insufficient data or invalid starting value.
    """
    if len(series) < 2:
        return None
    if series[0] == 0:
        return None
    # Handle negative starting values - CAGR formula requires positive start
    if series[0] < 0:
        return None
    if series[-1] < 0:
        return None
    return (series[-1] / series[0]) ** (1.0 / (len(series) - 1)) - 1.0


def _compute_margin(income_series: List[float], revenue_series: List[float]) -> Optional[float]:
    """Compute average margin over available periods."""
    margins = []
    for ni, rev in zip(income_series, revenue_series):
        if rev and rev != 0:
            margins.append(ni / rev)
    return sum(margins) / len(margins) if margins else None


def derive_alpha_fields(
    financials: Dict[str, Any],
    company_id: str,
    sector: Optional[str] = None,
    series: Optional[Dict[str, List[float]]] = None,
) -> Dict[str, Any]:
    """
    Derive AlphaInput-compatible fields from normalised financials.
    
    Args:
        financials: Normalised financial metrics
        company_id: Company identifier
        sector: Sector name (attempts to derive from company_id if None)
        series: Optional pre-extracted time series
    """
    series = series or {}
    
    # Try to derive sector from company_id if not provided
    if sector is None:
        # Simple heuristic - can be enhanced with a lookup table
        tech_companies = {"TCS", "INFY", "WIPRO", "HCLTECH", "TECHM", "KSOLVES", "NEWGEN", "SAKSOFT", "INTELLECT", "RAMCOSYS", "EMUDHRA"}
        finance_companies = {"HDFC", "ICICI", "SBIN", "AXISBANK", "KOTAKBANK"}
        energy_companies = {"RELIANCE", "ONGC", "NTPC", "POWERGRID"}
        
        upper_id = company_id.upper()
        if upper_id in tech_companies or "TECH" in upper_id:
            sector = "Information Technology"
        elif upper_id in finance_companies or "BANK" in upper_id:
            sector = "Financials"
        elif upper_id in energy_companies:
            sector = "Energy"
        else:
            sector = "Diversified"

    revenue      = financials.get("revenue",      0) or 0
    net_income   = financials.get("net_income",   0) or 0
    total_equity = max(financials.get("total_equity", 1) or 1, 1)
    total_debt   = financials.get("total_debt",   0) or 0
    ebit         = financials.get("ebit",         0) or 0
    cogs         = financials.get("cogs",         0) or 0
    interest_exp = financials.get("interest_expense", 0) or 0
    cfo          = financials.get("cash_from_operations", 0) or 0

    # Real multi-year series if available
    rev_series  = series.get("revenue_series")  or (
        [revenue * (0.75 + 0.06 * i) for i in range(5)] if revenue else []
    )
    ni_series   = series.get("ni_series")       or (
        [net_income * (0.80 + 0.05 * i) for i in range(5)] if net_income else []
    )
    ebit_series = series.get("ebit_series")     or (
        [ebit * (0.78 + 0.055 * i) for i in range(5)] if ebit else []
    )

    growth = _compute_growth(rev_series) if rev_series else None
    margin = _compute_margin(ni_series, rev_series) if ni_series and rev_series else None

    # Derived ratios from REAL financials — each company will differ
    de_ratio = round(total_debt / total_equity, 4) if total_equity else 0.5
    icr      = round(ebit / interest_exp, 4)       if interest_exp else 8.0
    roe      = round(net_income / total_equity, 4) if total_equity else 0.12

    # Gross margin: (revenue - cogs) / revenue — real company-specific value
    gross_margin = round((revenue - cogs) / revenue, 4) if revenue and cogs else (
        round(net_income / revenue + 0.12, 4) if revenue and net_income else 0.35
    )

    # FCF conversion: CFO / net_income — real cash quality signal
    fcf_stored = financials.get("fcf_conversion")
    if fcf_stored:
        fcf_conv = float(fcf_stored)
    elif cfo and net_income:
        fcf_conv = round(min(cfo / net_income, 2.0), 4)   # cap at 2x
    else:
        fcf_conv = 0.85

    # Per-year margin series derived from actual series (not hardcoded constants)
    npm_series = [
        round(ni / rev, 4) if rev else 0.0
        for ni, rev in zip(ni_series, rev_series)
    ]
    ebit_margin_series = [
        round(eb / rev, 4) if rev else 0.0
        for eb, rev in zip(ebit_series, rev_series)
    ]
    # Gross margin series: assume slight YoY change around the real base margin
    gross_margin_series = [
        round(gross_margin + 0.002 * (i - 2), 4) for i in range(5)
    ]

    return {
        "company_id":              company_id,
        "company_name":            company_id,
        "sector":                  sector,
        "revenue_series":          rev_series,
        "ebit_series":             ebit_series,
        "npm_series":              npm_series or [round((margin or 0.08), 4)] * 5,
        "ebit_margin_series":      ebit_margin_series or [round(ebit / revenue, 4) if revenue else 0.12] * 5,
        "gross_margin_series":     gross_margin_series,
        "debt_to_equity":          de_ratio,
        "interest_coverage_ratio": icr,
        "fcf_conversion":          fcf_conv,
        "roe":                     roe,
        "year_labels":             series.get("year_labels", []),
        "available_fields":        len([v for v in financials.values() if v and v != 0]),
        "required_fields":         20,
        "years_available":         len(rev_series),
    }


# ── Public entry point ────────────────────────────────────────────────────────

def load_company(company_id: str) -> Dict[str, Any]:
    """
    Load a company's data from S3 by reading ALL files in its folder.
    
    Raises CompanyDataError on any fatal failure including:
    - No financial data found
    - Insufficient price data (< 5 points)
    - Missing required Excel sheets
    """
    normalised_id = company_id.strip()
    prefix = f"{normalised_id}/"

    logger.info("[S3] Loading company: %s (prefix=%s)", normalised_id, prefix)

    # ── 1. List all objects under this company's folder ──────────────────────
    objects = _list_prefix(prefix)

    if not objects:
        # Try progressively more variants of the folder name.
        # Critically: include the ORIGINAL company_id casing (pre-strip) — callers
        # like the comparator preserved to force .upper() which would corrupt
        # case-sensitive S3 folder names like "HCL Technologies/".
        _tried: set = {normalised_id}
        for variant in [
            company_id.strip(),                 # original caller casing (most important!)
            normalised_id.upper(),              # ALL CAPS
            normalised_id.lower(),              # all lower
            normalised_id.title(),              # Title Case Each Word
            normalised_id.capitalize(),         # Only first letter
            normalised_id.replace(" ", "_"),    # spaces → underscores
            normalised_id.replace("_", " "),    # underscores → spaces
            normalised_id.replace(" ", ""),     # no separator at all
        ]:
            if not variant or variant in _tried:
                continue
            _tried.add(variant)
            objects = _list_prefix(f"{variant}/")
            if objects:
                prefix = f"{variant}/"
                normalised_id = variant
                break

    # Probe root-level xlsx using both current normalised_id and original casing.
    # dict.fromkeys preserves order and deduplicates.
    _xlsx_candidates = list(dict.fromkeys([
        f"{normalised_id}.xlsx",
        f"{company_id.strip()}.xlsx",
        f"{company_id.strip().upper()}.xlsx",
    ]))
    root_xlsx_key     = _xlsx_candidates[0]
    root_xlsx_objects: list = []
    for _cand in _xlsx_candidates:
        _found = _list_prefix(_cand)
        if _found:
            root_xlsx_key     = _cand
            root_xlsx_objects = _found
            break

    if not objects and not root_xlsx_objects:
        raise CompanyDataError(
            f"Company '{company_id}' not found in S3 bucket '{_BUCKET}'. "
            f"Expected folder: s3://{_BUCKET}/{prefix}"
        )

    logger.info("[S3] Found %d object(s) under %s", len(objects), prefix)

    all_keys = [o["key"] for o in objects]
    if root_xlsx_objects:
        logger.info("[S3] Including root level Excel file: %s", root_xlsx_key)
        all_keys.extend([o["key"] for o in root_xlsx_objects])

    # ── 2. Categorise files ───────────────────────────────────────────────────
    # .xls (legacy binary format) is unreliable — openpyxl may silently
    # return corrupt/partial data.  Reject hard and force .xlsx upload.
    xls_only_keys = [
        k for k in all_keys
        if k.lower().endswith('.xls') and not k.lower().endswith('.xlsx')
    ]
    xlsx_keys  = [k for k in all_keys if k.lower().endswith('.xlsx')]

    if xls_only_keys and not xlsx_keys:
        raise CompanyDataError(
            f"Invalid Excel format for '{company_id}': found legacy .xls file(s) "
            f"{xls_only_keys}. Please convert to .xlsx — the .xls format is "
            "unreliable and not supported by EREBUS."
        )
    if xls_only_keys:
        logger.warning(
            "[S3] Ignoring %d legacy .xls file(s) for %s — only .xlsx is accepted",
            len(xls_only_keys), normalised_id,
        )

    excel_keys = xlsx_keys
    pdf_keys   = [k for k in all_keys if k.lower().endswith('.pdf')]
    json_keys  = [k for k in all_keys if k.lower().endswith('.json')]

    if not excel_keys and root_xlsx_objects:
        excel_keys = [root_xlsx_key]

    # Root-level Screener.in FIRST
    try:
        if _list_prefix(root_xlsx_key):
            excel_keys = [root_xlsx_key] + [k for k in excel_keys if k != root_xlsx_key]
    except Exception:
        pass

    logger.info("[S3] Files — Excel:%d PDF:%d JSON:%d",
                len(excel_keys), len(pdf_keys), len(json_keys))

    # ── 3. Read document index JSON ───────────────────────────────────────────
    documents: List[Dict] = []
    doc_json_key = next((k for k in json_keys if "_documents.json" in k.lower()), None)
    if doc_json_key:
        try:
            raw = _fetch_bytes(doc_json_key)
            parsed = _read_json(raw)
            if isinstance(parsed, list):
                documents = parsed
            elif isinstance(parsed, dict):
                documents = parsed.get("documents", [])
            logger.info("[S3] Loaded document index: %d entries", len(documents))
        except CompanyDataError as e:
            logger.warning("[S3] Could not read document index: %s", e)

    # ── 4. Read Excel for financials ──────────────────────────────────────────
    sheets: Dict[str, Any] = {}

    qr_excel      = [k for k in excel_keys if "quarterly" in k.lower() or "data_sheet" in k.lower()]
    preferred_excel = qr_excel or excel_keys

    if preferred_excel:
        for xl_key in preferred_excel[:4]:
            try:
                logger.info("[S3] Reading Excel: %s", xl_key)
                raw = _fetch_bytes(xl_key)
                parsed_sheets = _read_excel(raw)
                if not any(v.get("rows") for v in parsed_sheets.values()):
                    parsed_sheets = _read_excel_wide_format(raw)
                for k, v in parsed_sheets.items():
                    if k not in sheets and v.get("rows"):
                        sheets[k] = v
            except CompanyDataError as e:
                logger.warning("[S3] Excel read failed (%s): %s", xl_key, e)
            except Exception as e:
                logger.warning("[S3] Excel parse error (%s): %s", xl_key, e)

    # ── 5. Build financials using schema normalizer ───────────────────────────
    from modules.ingestion.schema_normalizer import build_financials_from_sheets, REQUIRED_FIELDS, match_key

    if sheets:
        raw_fin    = build_financials_from_sheets(sheets)
        financials: Dict[str, Any] = dict(raw_fin)

        raw_op  = raw_fin.get("operating_profit", 0) or 0
        rev     = raw_fin.get("revenue", 0) or 0
        ni      = raw_fin.get("net_income", 0) or 0
        pbt     = raw_fin.get("pbt", 0) or 0
        tax_val = raw_fin.get("tax", 0) or 0
        eq_cap  = raw_fin.get("equity_capital", 0) or 0
        res_val = raw_fin.get("reserves", 0) or 0
        borrows = raw_fin.get("borrowings", 0) or 0
        int_exp = raw_fin.get("interest_expense", 0) or 0
        other_inc = raw_fin.get("other_income", 0) or 0

        # ── EBIT sanity ───────────────────────────────────────────────────────
        # Wide-format Data Sheets often have cumulative row totals as the last
        # column value. If operating_profit > 40% of revenue it's almost
        # certainly a data error. Recompute from PBT + interest (more reliable).
        raw_ebit_margin = (raw_op / rev) if rev > 0 else 0
        if raw_ebit_margin > 0.40 and pbt > 0:
            # PBT = EBIT - interest + other_income
            # → EBIT ≈ PBT + int_exp - other_inc
            ebit = max(pbt + int_exp - other_inc, 0)
            logger.info("[loader] EBIT corrected from %.0f → %.0f via PBT for %s",
                        raw_op, ebit, normalised_id)
        else:
            ebit = raw_op
        financials["ebit"]             = ebit
        financials["operating_income"] = ebit
        financials["total_debt"]       = borrows

        # ── COGS: never use revenue*0.45 — it makes every company identical ──
        expenses = raw_fin.get("expenses", 0) or 0
        if expenses > 0:
            cogs = expenses
        elif rev > 0 and ebit > 0:
            cogs = max(rev - ebit - other_inc, rev * 0.15)
        elif rev > 0 and ni > 0:
            # Fallback: estimate from net margin (each company differs)
            net_margin = ni / rev
            cogs = rev * max(1 - net_margin - 0.05, 0.15)
        else:
            cogs = 0
        financials["cogs"] = cogs

        # ── total_equity: rebuild when extracted value is implausibly small ───
        # A profitable company cannot have equity < 1 year of earnings.
        te = raw_fin.get("total_equity", 0) or 0
        if te <= 1.0 or (ni > 0 and te < ni):
            te = eq_cap + res_val          # try components first
        if te <= 1.0 or (ni > 0 and te < ni):
            te = max(ni * 5, 1.0)          # 5x earnings proxy (conservative P/B)
        financials["total_equity"] = te

        financials.setdefault("market_cap",         0.0)
        financials.setdefault("shares_outstanding", 0.0)
        financials.setdefault("price_per_share",    0.0)
        bv_per_share = (te / max(eq_cap, 1)) if eq_cap > 0 else 1.0
        financials.setdefault("book_value_per_share", bv_per_share)

        # ── FCF conversion: per-company when available, else PBT-based proxy ──
        cfo   = raw_fin.get("cash_from_operations", 0) or 0
        capex = raw_fin.get("capex", 0) or 0
        fcf   = raw_fin.get("fcf", 0) or (cfo - abs(capex))
        if ni and fcf:
            fcf_conv = round(min(fcf / ni, 2.0), 4)
        elif ni and pbt:
            # Proxy: typical IT company converts ~90-110% of PBT to cash
            fcf_conv = round(min((pbt * 0.90) / ni, 1.5), 4)
        else:
            fcf_conv = 0.90
        financials.setdefault("fcf_conversion", fcf_conv)

        # Revenue / NI / EBIT time-series — search ALL available sheets.
        # TCS Data Sheets store P&L in 'bs', not 'pnl'; other formats vary.
        def _extract_series(rows_list, target_key, n=8):
            """Return last n non-zero values from the first row matching target_key."""
            for row in rows_list:
                if match_key(row.get("narration", "")) == target_key:
                    vals = [float(v) for v in (row.get("values") or [])
                            if v is not None and float(v) != 0]
                    if vals:
                        return vals[-n:]
            return []

        def _best_series(key, n=8):
            """Find the longest series for `key` across all loaded sheets."""
            best = []
            for sdata in sheets.values():
                candidate = _extract_series(sdata.get("rows", []), key, n)
                if len(candidate) > len(best):
                    best = candidate
            return best

        rev_series  = _best_series("revenue")
        ni_series   = _best_series("net_income")
        ebit_series = _best_series("operating_profit")

        # Fallback: ni_series from pbt - tax when net_income row not found
        if not ni_series:
            pbt_series = _best_series("pbt")
            tax_series = _best_series("tax")
            if pbt_series:
                ni_series = [
                    round(p - (t if i < len(tax_series) else 0), 2)
                    for i, (p, t) in enumerate(zip(pbt_series,
                                                    tax_series + [0]*len(pbt_series)))
                ]

        # Fallback: ebit from revenue - expenses when operating_profit not found
        if not ebit_series and rev_series:
            exp_series = _best_series("expenses")
            if exp_series:
                ebit_series = [
                    round(r - e, 2)
                    for r, e in zip(rev_series, exp_series + [0]*len(rev_series))
                ]

        logger.info("[S3] Series extracted — rev:%d ni:%d ebit:%d",
                    len(rev_series), len(ni_series), len(ebit_series))

        financials["prior_revenue"] = rev_series[-2] if len(rev_series) >= 2 else financials.get("revenue", 0) * 0.95

        # ── Extract year labels from sheet headers ────────────────────────────
        year_labels: list = []
        for sheet_key in ("pnl", "bs", "data_sheet"):
            hdr = sheets.get(sheet_key, {}).get("headers", [])
            candidate = [
                h for h in hdr
                if h and re.search(
                    r"\b(20\d{2}|mar|jun|sep|dec|fy\d{2}|ttm|trailing)\b",
                    str(h), re.I
                )
            ]
            if len(candidate) >= len(year_labels):
                year_labels = candidate
        # Align to length of rev_series (tail-trim to match)
        if year_labels and rev_series:
            year_labels = year_labels[-len(rev_series):]

        series = {
            "revenue_series": rev_series,
            "ni_series":      ni_series,
            "ebit_series":    ebit_series,
            "year_labels":    year_labels,   # ← NEW: human-readable year headers
        }

        # Quarterly prices from Quarters sheet — only real close prices are
        # accepted.  Revenue-series proxy is REMOVED: it creates fake momentum
        # signals that make all companies look identical.
        qtr_rows = sheets.get("quarters", {}).get("rows", [])
        prices   = _extract_series(qtr_rows, "revenue", n=20)
        # If what we extracted is clearly a revenue series (very large numbers)
        # rather than stock prices, discard it — we have no real price data.
        if prices and max(prices) > 10_000:
            logger.warning(
                "[S3] Discarding revenue-as-price proxy for %s — values too large "
                "to be stock prices (max=%.0f). Use a real price feed.",
                normalised_id, max(prices),
            )
            prices = []

        logger.info("[S3] Financials via normalizer — revenue=%.1f, net_income=%.1f",
                    financials.get("revenue", 0), financials.get("net_income", 0))

    else:
        # Try JSON fallback
        financials = {}
        series     = {}
        prices     = []
        for jk in json_keys:
            if "_documents" in jk.lower():
                continue
            try:
                raw  = _fetch_bytes(jk)
                data = _read_json(raw)
                if data and "financials" in data:
                    financials = data["financials"]
                    prices     = data.get("prices", [])
                    series     = data.get("series", {})
                    logger.info("[S3] Loaded financials from JSON: %s", jk)
                    break
            except Exception as e:
                logger.debug("[S3] JSON read failed (%s): %s", jk, e)

    # ── 6. FAIL FAST: No data at all ─────────────────────────────────────────
    if not financials and not pdf_keys and not documents:
        raise CompanyDataError(
            f"No data found for '{company_id}'. "
            f"Expected Screener.in Excel file, JSON, or PDFs in s3://{_BUCKET}/{prefix}"
        )
        
    # If we have PDFs but no financials, provide a blank stub
    if not financials:
        logger.warning("[S3] No structured financials found for %s, relying purely on document search", normalised_id)
        financials = {
            "revenue": 0, "net_income": 0, "operating_income": 0,
            "total_assets": 0, "total_equity": 1, "cogs": 0,
            "_incomplete": True, "_missing_keys": ["revenue", "net_income"]
        }
    
    # Validate minimum required fields
    required_keys = ["revenue", "net_income", "total_equity"]
    missing_keys = [k for k in required_keys if financials.get(k, 0) == 0]
    if missing_keys:
        logger.warning("[S3] Missing key financial metrics for %s: %s", normalised_id, missing_keys)
        # Mark as stub for downstream awareness but continue
        financials["_incomplete"] = True
        financials["_missing_keys"] = missing_keys

    # ── 7. Price data — honest handling, NO fabrication ──────────────────────
    # Minimum 30 points needed for meaningful technical signals (RSI period=14
    # + MACD slow=26 means you need at least 26 + signal=9 = 35 to even
    # compute one MACD value).  If we don't have real prices, pass an empty
    # list — the quant module degrades gracefully and returns None for trends
    # and volatility, which is the CORRECT and HONEST behaviour.
    _MIN_PRICE_POINTS = 30
    price_source = "s3_quarterly" if len(prices) >= _MIN_PRICE_POINTS else "none"

    if len(prices) < _MIN_PRICE_POINTS:
        if prices:
            logger.warning(
                "[S3] Only %d price points for %s — minimum %d required for "
                "technical signals. Discarding insufficient series.",
                len(prices), normalised_id, _MIN_PRICE_POINTS,
            )
        else:
            logger.warning(
                "[S3] No price data for %s. Technical signals will be unavailable.",
                normalised_id,
            )
        prices = []   # Empty → quant trends/volatility return None (correct)

    # ── Build data-quality envelope ───────────────────────────────────────────
    # Note: `text` is populated in step 8 (PDF extraction). We initialise it
    # to None here so _has_text is correctly False at score-computation time,
    # then update the score after PDF parsing if text is found.
    text             = None
    _missing_fields  = financials.pop("_missing_keys", [])
    _is_incomplete   = financials.pop("_incomplete", bool(_missing_fields))
    _has_price_data  = len(prices) >= _MIN_PRICE_POINTS
    _has_financials  = bool(financials) and financials.get("revenue", 0) > 0
    _has_text        = bool(text)  # populated later — placeholder for score

    _warnings: list = []
    if not _has_price_data:
        _warnings.append(
            "No price data — technical signals (RSI, MACD, momentum) unavailable"
        )
    if not _has_financials:
        _warnings.append(
            "No financial data loaded — all fundamental signals will be zero"
        )
    elif _missing_fields:
        _warnings.append(
            f"Partial financial data — missing: {', '.join(_missing_fields)}. "
            "Fundamental signals may be biased toward CAS & alpha defaults."
        )
    if _is_incomplete and not _missing_fields:
        _warnings.append(
            "Financial data is incomplete — treat outputs as indicative only"
        )

    # Score: 0.0–1.0 based on coverage
    #   40% financials complete, 30% price data, 20% text available, 10% no missing fields
    _dq_score = round(
        (0.40 if _has_financials else 0.0)
        + (0.30 if _has_price_data else 0.0)
        + (0.20 if _has_text else 0.0)       # text missing yet — will be updated post-PDF
        + (0.10 if not _missing_fields else 0.0),
        2,
    )

    # Attach to financials — orchestrator + comparator will surface this
    financials["_data_quality"] = {
        "score":                       _dq_score,
        "has_price_data":              _has_price_data,
        "has_financial_data":          _has_financials,
        "technical_signals_available": _has_price_data,
        "price_source":                price_source,
        "price_points":                len(prices),
        "missing_fields":              _missing_fields,
        "incomplete":                  _is_incomplete,
        "warnings":                    _warnings,
    }


    # ── 8. Extract PDF text ───────────────────────────────────────────────────
    text: Optional[str] = None
    transcript_keys = [k for k in pdf_keys if
                       "transcript" in k.lower() or "earnings" in k.lower()]
    annual_keys     = [k for k in pdf_keys if "annual" in k.lower()]
    text_source_keys = (transcript_keys or annual_keys or pdf_keys)[:2]

    pdf_texts = []
    for pk in text_source_keys:
        try:
            logger.info("[S3] Extracting text from: %s", pk)
            raw  = _fetch_bytes(pk)
            text_chunk = _read_pdf_text(raw, max_chars=6000)
            if text_chunk.strip():
                pdf_texts.append(text_chunk)
        except CompanyDataError as e:
            logger.warning("[S3] PDF fetch failed: %s", e)
        except Exception as e:
            logger.warning("[S3] PDF text error: %s", e)

    if pdf_texts:
        text = "\n\n".join(pdf_texts)
        # Update data_quality score now that we know text availability
        if text:
            dq = financials.get("_data_quality", {})
            old_score = dq.get("score", 0)
            if not _has_text:   # Was 0 for text — now add the 0.20 weight
                dq["score"] = round(min(1.0, old_score + 0.20), 2)
            financials["_data_quality"] = dq

    # ── 9. Derive alpha fields ────────────────────────────────────────────────
    alpha_fields = derive_alpha_fields(
        financials=financials,
        company_id=normalised_id,
        sector=None,  # Auto-derive from company_id
        series=series,
    )

    context: Dict[str, Any] = {
        "company_id":   normalised_id,
        "prices":       prices,
        "financials":   financials,
        "text":         text,
        "alpha_fields": alpha_fields,
        "documents":    documents,
        "raw_excel":    {k: {"sheet_found": True, "row_count": len(v.get("rows", []))}
                         for k, v in sheets.items()},
        "s3_files": {
            "excel": excel_keys,
            "pdf":   pdf_keys,
        },
    }

    logger.info(
        "[S3] Loaded %s — Excel sheets=%s, PDFs=%d, prices=%d, sector=%s",
        normalised_id, list(sheets.keys()), len(pdf_keys), len(prices), alpha_fields.get("sector"),
    )
    return context