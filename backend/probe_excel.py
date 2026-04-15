"""
Peek at TCS Excel files to understand their sheet structure.
"""
from dotenv import load_dotenv
load_dotenv('.env', override=True)
import os, io, boto3
from botocore.config import Config
import openpyxl

client = boto3.client('s3',
    aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
    aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
    region_name=os.getenv('AWS_REGION','ap-south-1').strip(),
    config=Config(connect_timeout=3, read_timeout=15,
                  retries={'max_attempts':2,'mode':'standard'}))

BUCKET = os.getenv('S3_BUCKET_NAME','itsector1')

targets = [
    "TCS.xlsx",                                            # root-level Screener.in
    "TCS/Quarterly_Results/2026_TCS-Data-Sheet-Q3FY26.xlsx",  # folder Excel
]

for key in targets:
    print(f"\n=== {key} ===")
    try:
        obj = client.get_object(Bucket=BUCKET, Key=key)
        raw = obj['Body'].read()
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sname in wb.sheetnames[:4]:
            ws = wb[sname]
            print(f"\n  [{sname}] {ws.max_row}r x {ws.max_column}c")
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i > 10: break
                cells = [str(c)[:20] for c in row if c is not None]
                if cells:
                    print(f"    r{i}: {cells}")
    except Exception as e:
        print(f"  ERROR: {e}")
